from io import BytesIO
from urllib.parse import urlencode
from collections import defaultdict

from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.http import HttpResponse
from django.shortcuts import render
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from apps.actividades.models import Actividad

from .forms import ReportesFiltroForm


TIPOS_REPORTE = {
    'general': 'Reporte General',
    'tipologia': 'Por Tipologia',
    'programa': 'Por Programa',
    'mensual': 'Mensual',
    'anual': 'Anual',
    'personalizado': 'Personalizado',
}


def _construir_datos_graficas(actividades):
    tipologia_labels = dict(Actividad.TIPOLOGIAS)
    modalidad_labels = dict(Actividad.MODALIDADES)
    meses_labels = dict(Actividad.MESES)
    meses_orden = [codigo for codigo, _ in Actividad.MESES]

    conteo_tipologia = defaultdict(int)
    conteo_modalidad = defaultdict(int)
    conteo_programa = defaultdict(int)

    mensual_actividades = {mes: 0 for mes in meses_orden}
    mensual_participantes = {mes: 0 for mes in meses_orden}
    mensual_horas = {mes: 0 for mes in meses_orden}

    for actividad in actividades:
        conteo_tipologia[actividad.tipologia] += 1
        conteo_modalidad[actividad.modalidad] += 1
        conteo_programa[actividad.programa] += 1

        mes_codigo = actividad.mes if actividad.mes in mensual_actividades else None
        if mes_codigo:
            mensual_actividades[mes_codigo] += 1
            mensual_participantes[mes_codigo] += actividad.numero_participantes or 0
            mensual_horas[mes_codigo] += actividad.horas_dedicadas or 0

    tipologias_presentes = [
        codigo
        for codigo, _ in Actividad.TIPOLOGIAS
        if conteo_tipologia.get(codigo, 0) > 0
    ]

    modalidades_presentes = [
        codigo
        for codigo, _ in Actividad.MODALIDADES
        if conteo_modalidad.get(codigo, 0) > 0
    ]

    programas_ordenados = sorted(
        conteo_programa.items(),
        key=lambda item: (-item[1], item[0].lower()),
    )
    top_programas = programas_ordenados[:8]
    resto_programas = programas_ordenados[8:]
    otros_programas = sum(cantidad for _, cantidad in resto_programas)

    programas_labels = [nombre for nombre, _ in top_programas]
    programas_values = [cantidad for _, cantidad in top_programas]
    if otros_programas:
        programas_labels.append('Otros')
        programas_values.append(otros_programas)

    return {
        'tipologias': {
            'labels': [tipologia_labels[codigo] for codigo in tipologias_presentes],
            'values': [conteo_tipologia[codigo] for codigo in tipologias_presentes],
        },
        'modalidades': {
            'labels': [modalidad_labels[codigo] for codigo in modalidades_presentes],
            'values': [conteo_modalidad[codigo] for codigo in modalidades_presentes],
        },
        'programas': {
            'labels': programas_labels,
            'values': programas_values,
        },
        'mensual': {
            'labels': [meses_labels[codigo] for codigo in meses_orden],
            'actividades': [mensual_actividades[codigo] for codigo in meses_orden],
            'participantes': [mensual_participantes[codigo] for codigo in meses_orden],
            'horas': [mensual_horas[codigo] for codigo in meses_orden],
        },
    }


def _parse_anio(valor):
    try:
        return int(valor)
    except (TypeError, ValueError):
        return None


def _filtrar_actividades(tipo_reporte, filtros):
    actividades = Actividad.objects.all()

    tipologia = filtros.get('tipologia')
    programa = filtros.get('programa')
    mes = filtros.get('mes')
    anio = filtros.get('anio')
    fecha_inicio = filtros.get('fecha_inicio')
    fecha_fin = filtros.get('fecha_fin')

    if isinstance(tipologia, str):
        tipologia = tipologia.strip()
    if isinstance(programa, str):
        programa = programa.strip()
    if isinstance(mes, str):
        mes = mes.strip()
    if isinstance(anio, str):
        anio = _parse_anio(anio.strip())
    elif anio is not None:
        anio = _parse_anio(anio)

    if isinstance(fecha_inicio, str):
        fecha_inicio = fecha_inicio.strip() or None
    if isinstance(fecha_fin, str):
        fecha_fin = fecha_fin.strip() or None

    if tipologia:
        actividades = actividades.filter(tipologia=tipologia)

    if programa:
        actividades = actividades.filter(programa=programa)

    if mes:
        actividades = actividades.filter(mes=mes)

    if anio:
        actividades = actividades.filter(fecha_inicio__year=anio)

    if fecha_inicio:
        actividades = actividades.filter(fecha_inicio__gte=fecha_inicio)
    if fecha_fin:
        actividades = actividades.filter(fecha_fin__lte=fecha_fin)

    return actividades


def _construir_filtros(request):
    return {
        'tipologia': request.GET.get('tipologia', ''),
        'programa': request.GET.get('programa', ''),
        'mes': request.GET.get('mes', ''),
        'anio': request.GET.get('anio', ''),
        'fecha_inicio': request.GET.get('fecha_inicio', ''),
        'fecha_fin': request.GET.get('fecha_fin', ''),
    }


def _querystring_excel(tipo_reporte, filtros):
    parametros = {'tipo': tipo_reporte}
    for clave, valor in filtros.items():
        if valor:
            parametros[clave] = valor
    return urlencode(parametros)


def _crear_libro_excel(tipo_reporte, actividades):
    libro = Workbook()
    hoja = libro.active
    hoja.title = 'Reporte'

    encabezados = [
        'Tipo de Reporte',
        'Nombre Actividad',
        'Programa',
        'Tipologia',
        'Modalidad',
        'Periodo',
        'Fecha Inicio',
        'Fecha Fin',
        'Participantes',
        'Horas',
    ]
    hoja.append(encabezados)

    encabezado_fill = PatternFill(fill_type='solid', fgColor='1F4E78')
    encabezado_font = Font(bold=True, color='FFFFFF')

    for indice_columna in range(1, len(encabezados) + 1):
        celda = hoja.cell(row=1, column=indice_columna)
        celda.fill = encabezado_fill
        celda.font = encabezado_font
        celda.alignment = Alignment(horizontal='center')

    for actividad in actividades:
        hoja.append([
            TIPOS_REPORTE.get(tipo_reporte, TIPOS_REPORTE['general']),
            actividad.nombre,
            actividad.programa,
            actividad.get_tipologia_display(),
            actividad.get_modalidad_display(),
            actividad.periodo,
            actividad.fecha_inicio.strftime('%Y-%m-%d') if actividad.fecha_inicio else '',
            actividad.fecha_fin.strftime('%Y-%m-%d') if actividad.fecha_fin else '',
            actividad.numero_participantes,
            actividad.horas_dedicadas,
        ])

    for columna in hoja.columns:
        longitud_maxima = max(len(str(celda.value)) if celda.value is not None else 0 for celda in columna)
        hoja.column_dimensions[columna[0].column_letter].width = min(max(longitud_maxima + 2, 12), 45)

    return libro


@login_required
def reportes(request):
    form = ReportesFiltroForm(request.GET or None, tipos_reporte=TIPOS_REPORTE)
    if form.is_valid():
        filtros = form.cleaned_data
    else:
        filtros = _construir_filtros(request)

    tipo_reporte = filtros.get('tipo') or 'general'
    if tipo_reporte not in TIPOS_REPORTE:
        tipo_reporte = 'general'

    actividades = _filtrar_actividades(tipo_reporte, filtros)
    totales = actividades.aggregate(
        total_participantes=Sum('numero_participantes'),
        total_horas=Sum('horas_dedicadas'),
    )
    datos_graficas = _construir_datos_graficas(actividades)

    anios_disponibles = sorted(
        {actividad.fecha_inicio.year for actividad in Actividad.objects.only('fecha_inicio')},
        reverse=True,
    )

    contexto = {
        'form': form,
        'anios_disponibles': anios_disponibles,
        'actividades': actividades,
        'total_actividades': actividades.count(),
        'total_participantes': totales['total_participantes'] or 0,
        'total_horas': totales['total_horas'] or 0,
        'querystring_excel': _querystring_excel(tipo_reporte, filtros),
        'datos_graficas': datos_graficas,
    }
    return render(request, 'reportes/reportes.html', contexto)


@login_required
def descargar_reporte_excel(request):
    tipo_reporte = request.GET.get('tipo', 'general')
    if tipo_reporte not in TIPOS_REPORTE:
        tipo_reporte = 'general'

    filtros = _construir_filtros(request)
    actividades = _filtrar_actividades(tipo_reporte, filtros)

    libro = _crear_libro_excel(tipo_reporte, actividades)
    buffer = BytesIO()
    libro.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = (
        f'attachment; filename="reporte_{tipo_reporte}_{request.user.username}.xlsx"'
    )

    return response
